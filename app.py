# IMPORT LIBRARIES

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import json
import os
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")



#app settings

APP_TITLE = "ClientFlow"

DATA_FILE = "clients.json"

selected_client_email = None

# create data file if it doesn`t exist.

if not os.path.exists(DATA_FILE):
    with open(DATA_FILE, "w") as file:
        json.dump([], file)

# this part is responsible for window.


root = ctk.CTk()
root.title(APP_TITLE)
root.geometry("750x560")
root.resizable(True, True)
root.minsize(750, 560)
root.configure(bg="#F5F7FA")


#  this part for title label


title_label = ctk.CTkLabel(
    root,
    text="ClientFlow",
    font=("Arial", 28, "bold"),
    text_color="#111827"
)

title_label.pack(pady=(15, 5))

####top form

form_frame = ctk.CTkFrame(
    root,
    fg_color="white",
    corner_radius=15
)
form_frame.pack(pady=8, padx=20, fill="x", ipadx=6, ipady=4)

####this part for input clinet name 

name_label = ctk.CTkLabel(
    form_frame,
    text="Client Name",
    text_color="#374151"
)

name_label.grid(row=0, column=0, padx=10)

name_entry = ctk.CTkEntry(
    form_frame,
    width=160,
    height=40,
    font=("Arial", 12),
    corner_radius=10
)
name_entry.grid(row=1, column=0, padx=10, pady=(0, 10))

####is for input  phone number


phone_label = ctk.CTkLabel(
    form_frame,
    text="Phone Number",
    text_color="#374151"
)

phone_label.grid(row=0, column=1, padx=10)

phone_entry = ctk.CTkEntry(
    form_frame,
    width=160,
    height=40,
    font=("Arial", 12),
    corner_radius=10
)
phone_entry.grid(row=1, column=1, padx=10, pady=(0, 10))

####

email_label = ctk.CTkLabel(
    form_frame,
    text="Email",
    text_color="#374151"
)

email_label.grid(row=0, column=2, padx=10)

email_entry = ctk.CTkEntry(
    form_frame,
    width=160,
    height=40,
    font=("Arial", 12),
    corner_radius=10
)

email_entry.grid(row=1, column=2, padx=10, pady=(0, 10))

####for input follow _up date

date_entry = DateEntry(
    form_frame,
    width=20,
    background="#2563EB",
    foreground="white",
    borderwidth=2,
    date_pattern="dd/mm/yyyy",
    year=2026,
    month=5,
    day=20,
    font=("Arial", 12)
)

date_entry.grid(row=1, column=3, padx=10, pady=(0, 10))

####seve new clint data

def save_client():
    global selected_client_email
####get data from input fields

    client_name = name_entry.get()
    phone_number = phone_entry.get()
    followup_date = date_entry.get()
    email = email_entry.get()

#### chek if any field is empty

    if client_name == "" or phone_number == "" or email == "" or followup_date == "":
        messagebox.showwarning("Missing Data", "Please fill all fields")
        return
    
####create client data dictionary

    client_data = {
        "name": client_name,
        "phone": phone_number,
        "email": email,
        "date": followup_date,
        "status": "Pending"
    }

    with open(DATA_FILE, "r") as file:
        clients = json.load(file)

    if selected_client_email is None:
        clients.append(client_data)

    else:
        for client in clients:
            if client.get("email") == selected_client_email:
                client["name"] = client_name
                client["phone"] = phone_number
                client["email"] = email
                client["date"] = followup_date

        selected_client_email = None

    with open(DATA_FILE, "w") as file:
        json.dump(clients, file, indent=4)

    name_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    date_entry.delete(0, tk.END)
    

    messagebox.showinfo("Success", "Client saved successfully")

    load_clients()


####save button

save_button = ctk.CTkButton(
    form_frame,
    text="Save",
    width=180,
    height=40,
    fg_color="#2563EB",
    hover_color="#1D4ED8",
    text_color="white",
    corner_radius=10,
    cursor="hand2",
    command=save_client
)

save_button.grid(row=2, column=0, columnspan=4, pady=6)

#### search input field


search_frame = ctk.CTkFrame(
    root,
    fg_color="white",
    corner_radius=15
)
search_frame.pack(
    pady=6,
    padx=20,
    fill="x",
    ipadx=6,
    ipady=4
)

search_label = ctk.CTkLabel(
    search_frame,
    text="🔍 Search Client",
    text_color="#374151"
)

search_label.pack(side="left", padx=5)

search_entry = ctk.CTkEntry(
    search_frame,
    width=1300,
    height=34,
    font=("Arial", 12),
    corner_radius=10
)
search_entry.pack(side="left", padx=5)

search_entry.bind("<KeyRelease>", lambda event: search_clients())

# dashboard cards

cards_frame = ctk.CTkFrame(
    root,
    fg_color="transparent"
)

cards_frame.pack(fill="x", padx=20, pady=10)

total_card = ctk.CTkFrame(
    cards_frame,
    fg_color="white",
    corner_radius=15
)

total_card.pack(side="left", expand=True, fill="x", padx=5)

total_title = ctk.CTkLabel(
    total_card,
    text="Total Clients",
    text_color="#6B7280",
    font=("Arial", 12)
)

total_title.pack(pady=(3, 0))

total_value = ctk.CTkLabel(
    total_card,
    text="0",
    text_color="#111827",
    font=("Arial", 24, "bold")
)

total_value.pack(pady=(0, 3))

####pending clients card

pending_card = ctk.CTkFrame(
    cards_frame,
    fg_color="white",
    corner_radius=15
)

pending_card.pack(side="left", expand=True, fill="x", padx=5)

pending_title = ctk.CTkLabel(
    pending_card,
    text="Pending",
    text_color="#6B7280",
    font=("Arial", 12)
)

pending_title.pack(pady=(3, 0))

pending_value = ctk.CTkLabel(
    pending_card,
    text="0",
    text_color="#F59E0B",
    font=("Arial", 24, "bold")
)

pending_value.pack(pady=(0, 3))

####contacted clients card

contacted_card = ctk.CTkFrame(
    cards_frame,
    fg_color="white",
    corner_radius=15
)

contacted_card.pack(side="left", expand=True, fill="x", padx=5)

contacted_title = ctk.CTkLabel(
    contacted_card,
    text="Contacted",
    text_color="#6B7280",
    font=("Arial", 12)
)

contacted_title.pack(pady=(3, 0))

contacted_value = ctk.CTkLabel(
    contacted_card,
    text="0",
    text_color="#16A34A",
    font=("Arial", 24, "bold")
)

contacted_value.pack(pady=(0, 3))

#### this part is responsible for table design

style = ttk.Style()

style.configure(
    "Treeview",
    rowheight=32,
    font=("Arial", 11),
    background="white",
    fieldbackground="white"
)

style.configure(
    "Treeview.Heading",
    font=("Arial", 11, "bold"),
    background="#F3F4F6",
    foreground="#111827"
)

style.map(
    "Treeview",
    background=[("selected", "#DBEAFE")],
    foreground=[("selected", "#111827")]
)




####show client data in table

client_table = ttk.Treeview(
    root,
    columns=("Name", "Phone","Email", "Date" ,"Status"),
    show="headings",
    height=7
)

client_table.heading("Name", text="Client Name")
client_table.heading("Phone", text="Phone Number")
client_table.heading("Email", text="Email")
client_table.heading("Date", text="Follow-Up Date")
client_table.heading("Status", text="Status")

client_table.pack(
    pady=5,
    padx=20,
    fill="both"
   
)

client_table.column("Name", width=220)
client_table.column("Phone", width=180)
client_table.column("Email", width=220)
client_table.column("Date", width=180)
client_table.column("Status", width=120)

client_table.tag_configure(
    "contacted",
    background="#DCFCE7"
)

client_table.tag_configure(
    "pending",
    background="#FEF3C7"
)

####update dashboard 

def update_dashboard():
    with open(DATA_FILE, "r") as file:
        clients = json.load(file)

    total_count = len(clients)
    pending_count = 0
    contacted_count = 0

    for client in clients:
        status = client.get("status", "Pending")

        if status == "Pending":
            pending_count += 1

        if status == "Contacted":
            contacted_count += 1

    total_value.configure(text=str(total_count))
    pending_value.configure(text=str(pending_count))
    contacted_value.configure(text=str(contacted_count))

    

##### Show saved clients in the table

def load_clients():

    for row in client_table.get_children():
        client_table.delete(row)


    with open(DATA_FILE, "r") as file:
        clients = json.load(file)

    for client in clients:

        status = client.get("status", "Pending")

        tag_name = "contacted" if status == "Contacted" else "pending"

        client_table.insert(
            "",
            tk.END,
            values=(
                client["name"],
                client["phone"],
                client.get("email", ""),
                client["date"],
                status
         ),
         tags=(tag_name,)
    )

    update_dashboard()

load_clients()

#####search clients in the table based on name or phone number

def search_clients():
    search_text = search_entry.get().lower()

    for row in client_table.get_children():
        client_table.delete(row)

    with open(DATA_FILE, "r") as file:
        clients = json.load(file)

    for client in clients:
        if (
            search_text in client["name"].lower()
            or search_text in client["phone"]
        ):
            client_table.insert(
                "",
                tk.END,
                values=(
                    client["name"],
                    client["phone"],
                    client.get("email", ""),
                    client["date"],
                    client.get("status", "Pending")
                )
            )


####this is foer mark client as pending 


def mark_pending():
    selected_item = client_table.selection()

    if not selected_item:
        messagebox.showwarning("No Selection", "Please select a client first")
        return

    

    with open(DATA_FILE, "r") as file:
        clients = json.load(file)

    selected_values = client_table.item(selected_item[0])["values"]

    for client in clients:

        if (
            str(client["name"]) == str(selected_values[0])
            and str(client.get("email", "")) == str(selected_values[2])
    ):
            client["status"] = "Pending"

    with open(DATA_FILE, "w") as file:
        json.dump(clients, file, indent=4)

    load_clients()


##### this is for make client as contacted


def mark_contacted():
    selected_item = client_table.selection()

    if not selected_item:
        messagebox.showwarning("No Selection", "Please select a client first")
        return

    

    with open(DATA_FILE, "r") as file:
      clients = json.load(file)

    selected_values = client_table.item(selected_item[0])["values"]
    print("Selected values:", selected_values)

    for client in clients:
        print("Checking client:", client)

        if (
              str(client["name"]) == str(selected_values[0])
             and str(client.get("email", "")) == str(selected_values[2])
        ):
             client["status"] = "Contacted"

    with open(DATA_FILE, "w") as file:
        json.dump(clients, file, indent=4)

    load_clients()
 ##### is for delete client from table and data file 


def delete_client():
    selected_item = client_table.selection()

    if not selected_item:
        messagebox.showwarning("No Selection", "Please select a client first")
        return

    selected_values = client_table.item(selected_item[0])["values"]

    with open(DATA_FILE, "r") as file:
        clients = json.load(file)

    updated_clients = []

    for client in clients:
        if not (
            str(client["name"]) == str(selected_values[0])
            and str(client.get("email", "")) == str(selected_values[2])
        ):
            updated_clients.append(client)

    with open(DATA_FILE, "w") as file:
        json.dump(updated_clients, file, indent=4)

    client_table.delete(selected_item[0])

    messagebox.showinfo("Deleted", "Client deleted successfully")

    update_dashboard()

#### this is for export client data to csv file


def export_clients():
    with open(DATA_FILE, "r") as file:
        clients = json.load(file)

    if len(clients) == 0:
        messagebox.showwarning("No Data", "No clients to export")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clients"

    sheet.append(["Name", "Phone", "Email", "Follow-Up Date", "Status"])

    for client in clients:
        sheet.append([
            client["name"],
            client["phone"],
            client.get("email", ""),
            client["date"],
            client.get("status", "Pending")
        ])

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile="clients_export.xlsx"
    )

    if file_path == "":
        return

    workbook.save(file_path)

    messagebox.showinfo("Export Done", "Excel file saved successfully")

####

def import_clients():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")]
    )

    if file_path == "":
        return

    workbook = load_workbook(file_path)
    sheet = workbook.active

    with open(DATA_FILE, "r") as file:
        clients = json.load(file)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, phone, email, followup_date, status = row

        client_data = {
            "name": str(name),
            "phone": str(phone),
            "email": str(email),
            "date": str(followup_date),
            "status": str(status)
        }

        clients.append(client_data)

    with open(DATA_FILE, "w") as file:
        json.dump(clients, file, indent=4)

    load_clients()

    messagebox.showinfo("Import Done", "Excel file imported successfully")

####


#### here you can find create buttons for mark as contacted, pending, delete and export data


####contacted button


button_frame = ctk.CTkFrame(
    root,
    fg_color="white",
    corner_radius=15
)

button_frame.pack(
    pady=8,
    padx=20,
    fill="x",
    ipadx=6,
    ipady=4
)

contacted_button = ctk.CTkButton(
    button_frame,
    text="Mark Contacted",
    width=130,
    height=34,
    fg_color="#16A34A",
    hover_color="#15803D",
    text_color="white",
    corner_radius=8,
    cursor="hand2",
    command=mark_contacted
)

contacted_button.pack(side="left", padx=5)


####pending button


pending_button = ctk.CTkButton(
    button_frame,
    text="Mark Pending",
    width=130,
    height=38,
    fg_color="#F59E0B",
    hover_color="#D97706",
    text_color="white",
    corner_radius=8,
    cursor="hand2",
    command=mark_pending
)

pending_button.pack(side="left", padx=5)


#### delete button


delete_button = ctk.CTkButton(
    button_frame,
    text="Delete Client",
    width=130,
    height=38,
    fg_color="#DC2626",
    hover_color="#B91C1C",
    text_color="white",
    corner_radius=8,
    cursor="hand2",
    command=delete_client
)

delete_button.pack(side="left", padx=5)


#### export button


export_button = ctk.CTkButton(
    button_frame,
    text="Export Excel",
    width=130,
    height=38,
    fg_color="#7C3AED",
    hover_color="#6D28D9",
    text_color="white",
    corner_radius=8,
    cursor="hand2",
    command=export_clients
)

export_button.pack(side="left", padx=5)

import_button = ctk.CTkButton(
    button_frame,
    text="Import Excel",
    width=130,
    height=38,
    fg_color="#0F766E",
    hover_color="#115E59",
    text_color="white",
    corner_radius=8,
    cursor="hand2",
    command=import_clients
)

import_button.pack(side="left", padx=5)

####

def on_double_click(event):

    global selected_client_email

    selected_item = client_table.selection()

    if not selected_item:
        return

    values = client_table.item(selected_item[0])["values"]

    selected_client_email = values[2]

    name_entry.delete(0, tk.END)
    name_entry.insert(0, values[0])

    phone_entry.delete(0, tk.END)
    phone_entry.insert(0, values[1])

    email_entry.delete(0, tk.END)
    email_entry.insert(0, values[2])

    date_entry.delete(0, tk.END)
    date_entry.insert(0, values[3])

client_table.bind("<Double-1>", on_double_click)

root.mainloop()

